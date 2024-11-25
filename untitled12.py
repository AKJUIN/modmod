import streamlit as st
import pandas as pd
import io
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

# Define fields for extraction
FIELDS_CONFIG = {
    "Module Code and name": "below",
    "Module component": "next",
    "Problem identified?": "next",
    "Problem addressed?": "next",
    "Problem identified": "below",
    "Action taken": "below"
}

# Extraction functions
def extract_data_from_docx(file):
    data = {field: None for field in FIELDS_CONFIG.keys()}
    document = Document(file)

    for table in document.tables:
        for row_idx, row in enumerate(table.rows):
            cells = [cell.text.strip() for cell in row.cells]
            for field, method in FIELDS_CONFIG.items():
                if method == "next":
                    for i, cell_text in enumerate(cells):
                        if field.lower() in cell_text.lower():
                            if i + 1 < len(cells):
                                data[field] = cells[i + 1]
                            break
                elif method == "below":
                    for col_idx, cell in enumerate(cells):
                        if field.lower() in cell.lower():
                            if row_idx + 1 < len(table.rows):
                                data[field] = table.rows[row_idx + 1].cells[col_idx].text.strip()
                            break
    return data

def process_uploaded_files(uploaded_files):
    all_data = [extract_data_from_docx(file) for file in uploaded_files]
    return pd.DataFrame(all_data)

# Save extracted data to Excel
def save_data_to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extracted Data")
        worksheet = writer.sheets["Extracted Data"]

        for col in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_length + 2, 10)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    return output.getvalue()

# Comparison functions
def compare_spreadsheets(file1, file2):
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)
    key = "Module component"

    if key not in df1.columns or key not in df2.columns:
        combined = pd.DataFrame({"Error": ["Module component column missing in one or both files"]})
        return df1, df2, combined

    combined = pd.merge(df1, df2, on=key, how="outer", suffixes=("_File1", "_File2"))
    problem_field_file1 = "Problem identified?_File1"
    problem_field_file2 = "Problem identified?_File2"

    if problem_field_file1 in combined.columns and problem_field_file2 in combined.columns:
        combined["Highlight"] = combined.apply(
            lambda row: "Yes" if (
                isinstance(row[problem_field_file1], str) and row[problem_field_file1].strip().lower() in ["y", "yes"]
                and isinstance(row[problem_field_file2], str) and row[problem_field_file2].strip().lower() in ["y", "yes"]
            ) else "", axis=1
        )
    else:
        combined["Highlight"] = "Error: Missing comparison columns"

    return df1, df2, combined

# Save comparison results to Excel
def save_comparison_to_excel(df1, df2, combined):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df1.to_excel(writer, index=False, sheet_name="File 1 Data")
        df2.to_excel(writer, index=False, sheet_name="File 2 Data")
        combined.to_excel(writer, index=False, sheet_name="Comparison")

        workbook = writer.book
        for sheet_name in ["File 1 Data", "File 2 Data", "Comparison"]:
            worksheet = writer.sheets[sheet_name]

            for col in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_length + 2, 10)

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
                worksheet.row_dimensions[row[0].row].height = 15

        if "Comparison" in writer.sheets:
            comparison_sheet = writer.sheets["Comparison"]
            for row in comparison_sheet.iter_rows(min_row=2, max_row=comparison_sheet.max_row, max_col=comparison_sheet.max_column):
                if row[-1].value == "Yes":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    return output.getvalue()

# Streamlit app
st.title("Document Data Extractor, Analyzer, and Comparator")

menu = st.sidebar.radio("Navigation", ["Extract Data", "Analyze Data", "Compare Spreadsheets"])

if menu == "Extract Data":
    uploaded_files = st.file_uploader("Upload .docx files", type=["docx"], accept_multiple_files=True)
    if uploaded_files:
        df = process_uploaded_files(uploaded_files)
        st.dataframe(df)
        excel_data = save_data_to_excel(df)
        st.download_button("Download Extracted Data", data=excel_data, file_name="extracted_data.xlsx")

elif menu == "Analyze Data":
    analysis_file = st.file_uploader("Upload Extracted Data (Excel)", type=["xlsx"])
    if analysis_file:
        df = pd.read_excel(analysis_file)
        st.write(f"**Count of 'Yes' in 'Problem identified?': {df['Problem identified?'].str.contains('Yes|Y', case=False, na=False).sum()}**")

elif menu == "Compare Spreadsheets":
    file1 = st.file_uploader("Upload First Spreadsheet", type=["xlsx"], key="file1")
    file2 = st.file_uploader("Upload Second Spreadsheet", type=["xlsx"], key="file2")
    if file1 and file2:
        df1, df2, combined = compare_spreadsheets(file1, file2)
        st.dataframe(combined)
        excel_data = save_comparison_to_excel(df1, df2, combined)
        st.download_button("Download Comparison", data=excel_data, file_name="comparison.xlsx")
